"""
构建 template_extended.xlsx 的脚本。
把原始 BOL 模板(只有2行明细)扩展成 MAX_ITEM_ROWS 行明细,并做样式调整。
以后要改格式(比如某列对齐方式、字体、边框), 改这个脚本重新跑一遍即可,
不用手动改二进制模板文件。
 
用法: python3 build_template.py
输入: template.xlsx (原始模板, 只有2行明细样例)
输出: template_extended.xlsx (给 generate_bol.py 用)
"""
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Alignment
import copy
 
SOURCE_TEMPLATE = "template.xlsx"
OUTPUT_TEMPLATE = "template_extended.xlsx"
MAX_ITEM_ROWS = 8  # 明细行预留上限(实际生成时会按真实条数收紧,这个只是上限)
 
 
def main():
    wb = load_workbook(SOURCE_TEMPLATE)
    ws = wb.active
 
    # 1. 解除所有合并单元格(insert_rows 对合并区域处理有bug,必须先拆开)
    merges = [str(mc) for mc in ws.merged_cells.ranges]
    for m in merges:
        ws.unmerge_cells(m)
 
    # 2. 在第26行前插入新明细行, 原本只有24,25两行样例
    insert_at = 26
    delta = MAX_ITEM_ROWS - 2
    ws.insert_rows(insert_at, delta)
 
    # 3. 按偏移量重新计算并应用合并区域
    for m in merges:
        min_col, min_row, max_col, max_row = range_boundaries(m)
        if min_row >= insert_at:
            min_row += delta
            max_row += delta
        new_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
        ws.merge_cells(new_range)
 
    # 4. 新插入的明细行套用第25行的样式(边框/字体/行高)
    for new_row in range(insert_at, insert_at + delta):
        for col in range(1, 9):
            src = ws.cell(row=25, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst._style = copy.copy(src._style)
        ws.row_dimensions[new_row].height = ws.row_dimensions[25].height
 
    # 5. 清空所有明细行(24~24+MAX_ITEM_ROWS-1)里模板自带的示例数据, 只留样式
    # 注意: ws.cell(row,col,value=None) 不会清空已有值(openpyxl遇到value=None时会跳过赋值),
    #       必须显式用 .value = None 才能真正清空
    for r in range(24, 24 + MAX_ITEM_ROWS):
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None
 
    # 6. 格式微调: Pallet Count(G列) 居中
    for r in range(24, 24 + MAX_ITEM_ROWS):
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
 
    wb.save(OUTPUT_TEMPLATE)
    print(f"已生成 {OUTPUT_TEMPLATE}, 明细行上限 {MAX_ITEM_ROWS} 行")
 
 
if __name__ == "__main__":
    main()