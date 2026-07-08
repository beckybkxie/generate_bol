"""
BOL 批量生成脚本 v1
用法: python3 generate_bol.py <源数据文件.csv或.xlsx> <月份> <日期>
示例: python3 generate_bol.py 出入库记录导出.csv 7 6
 
逻辑:
1. 读取源表(腾讯文档导出的CSV或Excel), 按"派送时间"列筛选出指定月/日的记录
2. 按"送仓地址"(Ship To)分组 —— 同一个地址的多票货合并进一张BOL的明细表里
3. 每组套用 template_extended.xlsx, 填入动态字段, 输出一个xlsx
4. 用 LibreOffice 把每个 xlsx 转成 PDF
 
注意: SHIP FROM 固定用 SHIP_FROM_ADDRESS, Reference Number / ID Number 留空。
"""
import sys
import re
import os
import csv
import subprocess
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
 
TEMPLATE = "template_extended.xlsx"
ITEM_START_ROW = 24
MAX_ITEM_ROWS = 8  # 和 template_extended.xlsx 里预留的明细行数一致
OUT_DIR = "output"
 
# SHIP FROM 固定地址(用户确认: 所有BOL统一用这个发货地址)
SHIP_FROM_ADDRESS = "Equator\n218 Turnbull Canyon Rd, Hacienda Heights, CA 91745"
 
 
def match_date(val, month, day):
    if val is None:
        return False
    s = str(val)
    m = re.search(r"(\d{1,2})[/\-](\d{1,2})", s)
    if not m:
        return False
    return int(m.group(1)) == month and int(m.group(2)) == day
 
 
def normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"\s+", "", str(h))
 
 
def normalize_row(headers, values):
    norm_headers = [normalize_header(h) for h in headers]
    return dict(zip(norm_headers, values))
 
 
def load_csv_rows(path):
    # 腾讯文档导出的CSV编码不固定,依次尝试常见编码
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(path, encoding=enc, newline="") as f:
                reader = csv.reader(f)
                raw_headers = next(reader)
                rows = []
                for raw in reader:
                    d = normalize_row(raw_headers, raw)
                    if (d.get("柜号") or "").strip():
                        rows.append(d)
            return rows
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法用常见编码(utf-8/gbk等)读取 {path}, 请检查文件编码")
 
 
def load_xlsx_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = normalize_row(raw_headers, r)
        if d.get("柜号"):
            rows.append(d)
    return rows
 
 
def load_source_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return load_csv_rows(path)
    elif ext in (".xlsx", ".xlsm"):
        return load_xlsx_rows(path)
    else:
        raise ValueError(f"不支持的文件格式: {ext} (只支持 .csv / .xlsx)")
 
 
def to_number(val, default=0):
    if val is None or val == "":
        return default
    try:
        f = float(str(val).replace(",", "").strip())
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return default
 
 
def group_by_shipto(rows):
    groups = {}
    order = []
    for r in rows:
        key = (r.get("送仓地址") or "").strip()
        if not key:
            continue
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)
    return [(k, groups[k]) for k in order]
 
 
def fill_bol(items, ship_to, bol_number, bol_date, out_path):
    wb = load_workbook(TEMPLATE)
    ws = wb.active
 
    ws["H6"] = bol_date
    ws["G7"] = bol_number
 
    appt = items[0].get("派送时间") or ""
    ws["G8"] = str(appt)
 
    ws["A8"] = SHIP_FROM_ADDRESS
 
    ws["A15"] = ship_to
 
    n = min(len(items), MAX_ITEM_ROWS)
    for i in range(n):
        item = items[i]
        r = ITEM_START_ROW + i
        ws.cell(row=r, column=1, value=item.get("柜号") or "")
        ws.cell(row=r, column=2, value=item.get("参考号") or "")   # Reference Number, 源表暂无对应列
        ws.cell(row=r, column=3, value=item.get("ID号") or "")     # ID Number, 源表暂无对应列
        ws.cell(row=r, column=4, value=item.get("单号") or "")
        ctn = to_number(item.get("件数"))
        ws.cell(row=r, column=5, value=ctn)
        ws.cell(row=r, column=6, value=ctn)
        ws.cell(row=r, column=7, value=to_number(item.get("库存"), default=""))
        ws.cell(row=r, column=8, value=item.get("入库储位/location") or "")
 
    if len(items) > MAX_ITEM_ROWS:
        print(f"  ⚠️ 该BOL明细行数({len(items)})超过模板预留的{MAX_ITEM_ROWS}行, 已截断, 需要扩大MAX_ITEM_ROWS")
 
    extra = MAX_ITEM_ROWS - n
    if extra > 0:
        delete_start = ITEM_START_ROW + n
        merges = [str(mc) for mc in ws.merged_cells.ranges]
        for m in merges:
            ws.unmerge_cells(m)
        ws.delete_rows(delete_start, extra)
        for m in merges:
            min_col, min_row, max_col, max_row = range_boundaries(m)
            if min_row >= delete_start:
                min_row -= extra
                max_row -= extra
            new_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
            ws.merge_cells(new_range)
 
    wb.save(out_path)
 
 
import shutil
import platform
 
_SOFFICE_CACHE = {"path": None, "checked": False}
 
 
def find_soffice():
    if _SOFFICE_CACHE["checked"]:
        return _SOFFICE_CACHE["path"]
    _SOFFICE_CACHE["checked"] = True
 
    found = shutil.which("soffice") or shutil.which("soffice.exe")
    if found:
        _SOFFICE_CACHE["path"] = found
        return found
 
    if platform.system() == "Windows":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for c in candidates:
            if os.path.exists(c):
                _SOFFICE_CACHE["path"] = c
                return c
 
    return None
 
 
def convert_to_pdf(xlsx_path):
    soffice = find_soffice()
    if not soffice:
        print("  ⚠️ 找不到 soffice(LibreOffice), 跳过PDF转换, 只生成了xlsx。")
        print("     请确认已安装LibreOffice, 并把安装目录下的 program 文件夹加到PATH,")
        print(r"     或者确认路径是不是 C:\Program Files\LibreOffice\program\soffice.exe")
        return
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", OUT_DIR, xlsx_path],
        check=True, capture_output=True,
    )
 
 
def main():
    if len(sys.argv) != 4:
        print("用法: python3 generate_bol.py <源数据文件.csv或.xlsx> <月> <日>")
        sys.exit(1)
 
    source_path, month, day = sys.argv[1], int(sys.argv[2]), int(sys.argv[3])
    os.makedirs(OUT_DIR, exist_ok=True)
 
    rows = load_source_rows(source_path)
    if rows:
        print("识别到的字段(表头):", list(rows[0].keys()))
    target_rows = [r for r in rows if match_date(r.get("派送时间"), month, day)]
    print(f"筛选到 {month}/{day} 的记录: {len(target_rows)} 条")
 
    groups = group_by_shipto(target_rows)
    print(f"按送仓地址分组后, 共 {len(groups)} 张 BOL\n")
 
    today_str = date.today().strftime("%m/%d/%Y")
    seq = 1
    for ship_to, items in groups:
        bol_number = f"EBOL-{month:02d}{day:02d}-{seq:03d}"
        out_xlsx = os.path.join(OUT_DIR, f"{bol_number}.xlsx")
        fill_bol(items, ship_to, bol_number, today_str, out_xlsx)
        print(f"[{seq}] {bol_number}  明细{len(items)}条  -> {ship_to[:40]}")
        convert_to_pdf(out_xlsx)
        seq += 1
 
    print(f"\n完成, 输出目录: {OUT_DIR}/")
 
 
if __name__ == "__main__":
    main()